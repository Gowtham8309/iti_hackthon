from __future__ import annotations

import io
from datetime import date, datetime, time, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.database_models import AnomalyFlag, AttendanceEvent, Industry, Student


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "value"):
        return str(value.value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _header_style() -> tuple[Font, PatternFill, Alignment]:
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _apply_header_row(ws, headers: list[str]) -> None:
    font, fill, align = _header_style()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
    ws.row_dimensions[1].height = 28


def _auto_column_widths(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _today_window_utc(d: date) -> tuple[datetime, datetime]:
    start = datetime.combine(d, time.min, tzinfo=timezone.utc)
    end = datetime.combine(d, time.max, tzinfo=timezone.utc)
    return start, end


def export_attendance_events_excel(
    db: Session,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    student_id: int | None = None,
    status: str | None = None,
    limit: int = 5000,
) -> bytes:
    """
    Exports attendance events to an Excel (.xlsx) workbook and returns raw bytes.

    Filters: date range, student_id, status.
    """
    stmt = select(AttendanceEvent)

    if start_date:
        stmt = stmt.where(AttendanceEvent.event_time >= datetime.combine(start_date, time.min, tzinfo=timezone.utc))
    if end_date:
        stmt = stmt.where(AttendanceEvent.event_time <= datetime.combine(end_date, time.max, tzinfo=timezone.utc))
    if student_id is not None:
        stmt = stmt.where(AttendanceEvent.student_id == student_id)
    if status:
        stmt = stmt.where(AttendanceEvent.status == status.strip().lower())

    stmt = stmt.order_by(AttendanceEvent.event_time.desc(), AttendanceEvent.id.desc()).limit(limit)

    events = db.execute(stmt).scalars().all()

    student_cache: dict[int, Student] = {}
    industry_cache: dict[int, Industry] = {}

    def _get_student(sid: int) -> Student | None:
        if sid not in student_cache:
            student_cache[sid] = db.get(Student, sid)
        return student_cache[sid]

    def _get_industry(iid: int) -> Industry | None:
        if iid not in industry_cache:
            industry_cache[iid] = db.get(Industry, iid)
        return industry_cache[iid]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Events"

    headers = [
        "Event ID", "Event Type", "Status", "Event Time",
        "Student ID", "Student Code", "Student Name", "ITI Name", "Trade", "Batch",
        "Industry ID", "Industry Code", "Industry Name", "District", "City",
        "Latitude", "Longitude", "GPS Accuracy (m)", "Distance from Industry (m)",
        "Face Match Score", "Liveness Score",
        "Device ID", "IP Address", "Decision Reason",
        "Created At",
    ]

    _apply_header_row(ws, headers)

    for row_idx, event in enumerate(events, start=2):
        student = _get_student(event.student_id) if event.student_id else None
        industry = _get_industry(event.industry_id) if event.industry_id else None

        ws.append([
            event.id,
            _safe_str(event.event_type),
            _safe_str(event.status),
            _safe_str(event.event_time),
            event.student_id,
            student.student_code if student else "",
            student.full_name if student else "",
            student.iti_name if student else "",
            student.trade if student else "",
            student.batch if student else "",
            event.industry_id,
            industry.industry_code if industry else "",
            industry.name if industry else "",
            industry.district if industry else "",
            industry.city if industry else "",
            event.latitude,
            event.longitude,
            event.gps_accuracy_meters,
            event.distance_from_industry_meters,
            event.face_match_score,
            event.liveness_score,
            _safe_str(event.device_id),
            _safe_str(event.ip_address),
            _safe_str(event.decision_reason),
            _safe_str(event.created_at),
        ])

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_anomalies_excel(
    db: Session,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    student_id: int | None = None,
    severity: str | None = None,
    review_status: str | None = None,
    limit: int = 5000,
) -> bytes:
    """
    Exports anomaly flags to an Excel (.xlsx) workbook and returns raw bytes.
    """
    stmt = select(AnomalyFlag)

    if start_date:
        stmt = stmt.where(AnomalyFlag.created_at >= datetime.combine(start_date, time.min, tzinfo=timezone.utc))
    if end_date:
        stmt = stmt.where(AnomalyFlag.created_at <= datetime.combine(end_date, time.max, tzinfo=timezone.utc))
    if student_id is not None:
        stmt = stmt.where(AnomalyFlag.student_id == student_id)
    if severity:
        stmt = stmt.where(AnomalyFlag.severity == severity.strip().lower())
    if review_status:
        stmt = stmt.where(AnomalyFlag.review_status == review_status.strip().lower())

    stmt = stmt.order_by(AnomalyFlag.created_at.desc(), AnomalyFlag.id.desc()).limit(limit)

    anomalies = db.execute(stmt).scalars().all()

    student_cache: dict[int, Student] = {}

    def _get_student(sid: int) -> Student | None:
        if sid not in student_cache:
            student_cache[sid] = db.get(Student, sid)
        return student_cache[sid]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomaly Flags"

    headers = [
        "Anomaly ID", "Attendance Event ID",
        "Student ID", "Student Code", "Student Name", "ITI Name", "Trade", "Batch",
        "Anomaly Type", "Severity", "Risk Score", "Explanation",
        "Review Status", "Reviewed By (User ID)", "Reviewed At", "Review Notes",
        "Created At",
    ]

    _apply_header_row(ws, headers)

    for anomaly in anomalies:
        student = _get_student(anomaly.student_id) if anomaly.student_id else None

        ws.append([
            anomaly.id,
            anomaly.attendance_event_id,
            anomaly.student_id,
            student.student_code if student else "",
            student.full_name if student else "",
            student.iti_name if student else "",
            student.trade if student else "",
            student.batch if student else "",
            _safe_str(anomaly.anomaly_type),
            _safe_str(anomaly.severity),
            anomaly.risk_score,
            _safe_str(anomaly.explanation),
            _safe_str(anomaly.review_status),
            anomaly.reviewed_by_user_id,
            _safe_str(anomaly.reviewed_at),
            _safe_str(anomaly.review_notes),
            _safe_str(anomaly.created_at),
        ])

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
